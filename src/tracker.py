"""
CV tracking: saves PDF copies and updates Excel log for every sent CV.

Folder structure (under data/CVs enviados/):
  CVs/                  <- PDF copies of every sent CV
  Seguimiento Jobs.xlsx <- Excel tracker with all applications
"""
from __future__ import annotations

import os
import re
import shutil
import subprocess
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

TRACKER_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'CVs enviados')
CVS_DIR = os.path.join(TRACKER_DIR, 'CVs')
EXCEL_PATH = os.path.join(TRACKER_DIR, 'Seguimiento Jobs.xlsx')

HEADERS = [
    'Fecha',
    'Nombre de la posición',
    'Empresa',
    'Descripción',
    'Requerimientos',
    'Sueldo',
    'Link de la posición',
    'Ruta al CV',
]

COL_WIDTHS = [14, 35, 25, 60, 50, 18, 45, 55]


def _ensure_dirs():
    os.makedirs(CVS_DIR, exist_ok=True)


def _to_pdf(docx_path: str) -> str:
    """Convert DOCX to PDF using LibreOffice. Falls back to DOCX copy if unavailable."""
    try:
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'pdf', '--outdir', CVS_DIR, docx_path],
            capture_output=True, timeout=60
        )
        if result.returncode == 0:
            pdf_name = os.path.splitext(os.path.basename(docx_path))[0] + '.pdf'
            pdf_path = os.path.join(CVS_DIR, pdf_name)
            if os.path.exists(pdf_path):
                return pdf_path
    except Exception:
        pass
    # Fallback: copy DOCX as-is
    dest = os.path.join(CVS_DIR, os.path.basename(docx_path))
    shutil.copy2(docx_path, dest)
    return dest


def _extract_requirements(description: str) -> str:
    """Extract requirements section from job description."""
    if not description:
        return ''
    patterns = [
        r'(?:requirements?|qualifications?|what we(?:\'re| are) looking for|you (?:have|bring)|your profile'
        r'|requisitos?|perfil|experiencia requerida)[:\s]*\n((?:[-•*]\s*.+\n?){1,8})',
    ]
    for pattern in patterns:
        m = re.search(pattern, description, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:300]
    # Fallback: return first 200 chars of description
    return ''


def _extract_salary(description: str) -> str:
    """Try to extract salary/compensation info from description."""
    if not description:
        return ''
    patterns = [
        r'(?:salary|compensation|package|gehalt|salario)[:\s]*([€$£]?\s*[\d,\.]+(?:\s*[-–]\s*[\d,\.]+)?(?:\s*[kKmM])?(?:\s*(?:EUR|USD|GBP|ATS))?)',
        r'([€$£]\s*[\d,\.]+(?:\s*[-–]\s*[\d,\.]+)?\s*(?:per year|p\.a\.|annual|k)?)',
        r'([\d,\.]+\s*[-–]\s*[\d,\.]+\s*(?:EUR|USD|K|k))',
    ]
    for pattern in patterns:
        m = re.search(pattern, description, re.IGNORECASE)
        if m:
            return m.group(1).strip()[:60]
    return ''


def _create_excel():
    """Create a new Excel file with styled headers."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Seguimiento'

    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'
    wb.save(EXCEL_PATH)
    return wb


def record_sent_cv(job: dict, cv_docx_path: str) -> str:
    """
    Save a PDF copy of the CV and add a row to the Excel tracker.
    Returns the path to the saved PDF (or DOCX if conversion failed).
    """
    _ensure_dirs()

    # Save PDF copy
    pdf_path = _to_pdf(cv_docx_path)

    # Load or create Excel
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        _create_excel()
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

    description = job.get('description', '')
    row = [
        datetime.now().strftime('%d/%m/%Y'),
        job.get('title', ''),
        job.get('company', ''),
        (description[:500] + '...') if len(description) > 500 else description,
        _extract_requirements(description),
        _extract_salary(description),
        job.get('url', ''),
        pdf_path,
    ]

    ws.append(row)

    # Style data row
    data_row = ws.max_row
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=data_row, column=col_idx)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if data_row % 2 == 0:
            cell.fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')

    ws.row_dimensions[data_row].height = 60

    wb.save(EXCEL_PATH)
    print(f"[tracker] Recorded: {job['title']} @ {job['company']} | PDF: {pdf_path}")
    return pdf_path
